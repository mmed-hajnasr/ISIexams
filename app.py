import eel
import os
import pickle
import zipfile
import tempfile
import csv
import traceback
from seances import Seances, Seance
from enseignants import Enseignants
from configuration import Configuration
from assignements import Assignements
from datetime import datetime
from pdf_generation.surveillance_report import create_surveillance_report, create_enseignant_emploi

# Initialize Eel with the web directory
eel.init('web')

def convert_date_format(date_str, from_format='%Y-%m-%d', to_format='%d/%m/%Y'):
    """Convert date format between different formats"""
    try:
        date_obj = datetime.strptime(date_str, from_format)
        return date_obj.strftime(to_format)
    except ValueError:
        return date_str  # Return original if conversion fails

def save_current_state():
    """Save current application state to temporary files"""
    global seances_data, enseignants_data, configuration_data, assignements_data
    try:
        os.makedirs('data', exist_ok=True)
        
        # Save all application state in a single pickle file
        app_state = {
            'seances_data': seances_data,
            'enseignants_data': enseignants_data,
            'configuration_data': configuration_data,
            'assignements_data': assignements_data
        }
        
        with open('data/current_state.pkl', 'wb') as f:
            pickle.dump(app_state, f)
        
        # Also save configuration separately for backwards compatibility
        if configuration_data:
            with open('data/configuration.pkl', 'wb') as f:
                pickle.dump(configuration_data, f)
            
    except Exception as e:
        print(f"Warning: Could not save current state: {e}")

def load_current_state():
    """Load complete application state from temporary file"""
    try:
        if os.path.exists('data/current_state.pkl'):
            with open('data/current_state.pkl', 'rb') as f:
                app_state = pickle.load(f)
                
                # Check if it's the new format (dict with all data) or old format (just seances)
                if isinstance(app_state, dict) and 'seances_data' in app_state:
                    # Ensure all expected keys exist for backwards compatibility
                    if 'assignements_data' not in app_state:
                        app_state['assignements_data'] = None
                    return app_state
                else:
                    # Old format - just seances data
                    return {
                        'seances_data': app_state, 
                        'enseignants_data': None, 
                        'configuration_data': None,
                        'assignements_data': None
                    }
    except Exception as e:
        print(f"Warning: Could not load current state: {e}")
    return None

def load_configuration_state():
    """Load configuration state from temporary file"""
    try:
        if os.path.exists('data/configuration.pkl'):
            with open('data/configuration.pkl', 'rb') as f:
                return pickle.load(f)
    except Exception as e:
        print(f"Warning: Could not load configuration state: {e}")
    return Configuration()  # Return default configuration if none exists

# Global variables to store application state
seances_data = None
enseignants_data = None
configuration_data = None
assignements_data = None

@eel.expose
def initialize_app():
    """Initialize the application, loading previous state if available"""
    global seances_data, enseignants_data, configuration_data, assignements_data
    
    try:
        # Try to load previous state first
        app_state = load_current_state()
        
        if app_state:
            # Load from saved state
            seances_data = app_state.get('seances_data') or Seances()
            enseignants_data = app_state.get('enseignants_data')
            configuration_data = app_state.get('configuration_data')
            assignements_data = app_state.get('assignements_data')
            
        else:
            # Initialize with empty data
            seances_data = Seances()
            enseignants_data = None
            configuration_data = None
            assignements_data = None
        
        # Load configuration if not loaded from state
        if not configuration_data:
            configuration_data = load_configuration_state()
            
        # Load enseignants if not loaded from state
        if not enseignants_data:
            enseignants_original_path = 'data/enseignants.csv'
            
            if os.path.exists(enseignants_original_path):
                # Load from original import file
                enseignants_data = Enseignants.from_csv(enseignants_original_path)
            else:
                enseignants_data = Enseignants()
        
        # Always save the current state to ensure enseignants_data is persisted in pickle
        save_current_state()
            
        return {
            'success': True,
            'seances_info': {
                'semester': seances_data.semester,
                'exam_type': seances_data.exam_type,
                'session': seances_data.session,
                'dates': seances_data.dates,
                'total_sessions': sum(len(sessions) for sessions in seances_data.date_seances.values())
            },
            'enseignants_count': len(enseignants_data.enseignants_list)
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_available_dates():
    """Get all available exam dates"""
    global seances_data
    if seances_data:
        return seances_data.dates
    return []

@eel.expose
def get_seances_for_date(date):
    """Get all seances for a specific date"""
    global seances_data
    if not seances_data or date not in seances_data.date_seances:
        return []
    
    seances_list = []
    for i, seance in enumerate(seances_data.date_seances[date]):
        seances_list.append({
            'name': seances_data.get_seance_name(date, i),
            'h_debut': seance.h_debut,
            'h_fin': seance.h_fin,
            'salles': list(seance.salles),
            'enseignants': list(seance.responsables),
            'index': i
        })
    return seances_list

@eel.expose
def add_seance_to_date(date, h_debut, h_fin, salles=None, enseignants=None):
    """Add a new seance to a specific date"""
    global seances_data
    
    try:
        if not seances_data:
            seances_data = Seances()
            
        # Create new seance
        new_seance = Seance(h_debut=h_debut, h_fin=h_fin)
        
        # Add rooms if provided
        if salles:
            for salle in salles:
                if salle.strip():
                    new_seance.add_salle(salle.strip())
        
        # Add teachers if provided
        if enseignants:
            for enseignant in enseignants:
                if isinstance(enseignant, str) and enseignant.strip():
                    # Try to convert to int if it's a numeric string
                    try:
                        enseignant_code = int(enseignant.strip())
                        new_seance.add_enseignant(enseignant_code)
                    except ValueError:
                        # If it's not numeric, add as string (convert to int later if needed)
                        new_seance.add_enseignant(int(enseignant.strip()))
                elif isinstance(enseignant, int):
                    new_seance.add_enseignant(enseignant)
        
        # Add to seances data
        seances_data.add_seance(date, new_seance)
        
        # Sort seances by start time
        if date in seances_data.date_seances:
            seances_data.date_seances[date].sort(key=lambda s: s.h_debut)
        
        # Save current state
        save_current_state()
            
        return {
            'success': True,
            'message': f'Seance added successfully to {date}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def update_seance(date, index, h_debut, h_fin, salles, enseignants):
    """Update an existing seance"""
    global seances_data
    
    try:
        if not seances_data or date not in seances_data.date_seances:
            return {'success': False, 'error': 'Date not found'}
            
        if index >= len(seances_data.date_seances[date]):
            return {'success': False, 'error': 'Seance index out of range'}
            
        seance = seances_data.date_seances[date][index]
        
        # Update seance properties
        seance.h_debut = h_debut
        seance.h_fin = h_fin
        
        # Update salles
        seance.salles.clear()
        for salle in salles:
            seance.add_salle(salle.strip())
            
        # Update enseignants
        seance.enseignants.clear()
        for enseignant in enseignants:
            if isinstance(enseignant, str) and enseignant.isdigit():
                seance.add_enseignant(int(enseignant))
            elif isinstance(enseignant, int):
                seance.add_enseignant(enseignant)
                
        # Re-sort seances by start time
        seances_data.date_seances[date].sort(key=lambda s: s.h_debut)
        
        # Save current state
        save_current_state()
        
        return {
            'success': True,
            'message': 'Seance updated successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def delete_seance(date, index):
    """Delete a seance from a specific date"""
    global seances_data
    
    try:
        if not seances_data or date not in seances_data.date_seances:
            return {'success': False, 'error': 'Date not found'}
            
        if index >= len(seances_data.date_seances[date]):
            return {'success': False, 'error': 'Seance index out of range'}
            
        # Remove the seance
        del seances_data.date_seances[date][index]
        
        # If no seances left for this date, remove the date
        if not seances_data.date_seances[date]:
            del seances_data.date_seances[date]
            if date in seances_data.dates:
                seances_data.dates.remove(date)
        
        # Save current state
        save_current_state()
                
        return {
            'success': True,
            'message': 'Seance deleted successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def add_exam_date(date):
    """Add a new exam date (converts from YYYY-MM-DD to DD/MM/YYYY format)"""
    global seances_data
    
    try:
        if not seances_data:
            seances_data = Seances()
            
        # Convert date from HTML input format (YYYY-MM-DD) to DD/MM/YYYY
        formatted_date = convert_date_format(date, '%Y-%m-%d', '%d/%m/%Y')
            
        if formatted_date not in seances_data.dates:
            seances_data.dates.append(formatted_date)
            # Sort dates properly by converting to datetime objects
            seances_data.dates.sort(key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
            seances_data.date_seances[formatted_date] = []
            
            # Save current state
            save_current_state()
            
        return {
            'success': True,
            'message': f'Date {formatted_date} added successfully',
            'formatted_date': formatted_date
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def remove_exam_date(date):
    """Remove an exam date and all its seances"""
    global seances_data
    
    try:
        if not seances_data:
            return {'success': False, 'error': 'No seances data available'}
            
        if date in seances_data.dates:
            seances_data.dates.remove(date)
            
        if date in seances_data.date_seances:
            del seances_data.date_seances[date]
        
        # Save current state
        save_current_state()
            
        return {
            'success': True,
            'message': f'Date {date} removed successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_seances_summary():
    """Get a summary of all seances data"""
    global seances_data
    
    if not seances_data:
        return {
            'semester': None,
            'exam_type': None,
            'session': None,
            'dates': [],
            'total_sessions': 0,
            'seances_by_date': {}
        }
        
    return seances_data.to_dict()

@eel.expose
def import_seances_from_csv():
    """Import seances data from CSV file, replacing all existing data"""
    global seances_data
    
    try:
        seances_csv_path = 'data/seances.csv'
        if not os.path.exists(seances_csv_path):
            return {'success': False, 'error': f'CSV file not found: {seances_csv_path}'}
            
        # Load seances from CSV, replacing all existing data
        seances_data = Seances.from_csv(seances_csv_path)
        
        # Save the imported state
        save_current_state()
        
        return {
            'success': True,
            'message': 'Seances data imported successfully',
            'seances_info': {
                'semester': seances_data.semester,
                'exam_type': seances_data.exam_type,
                'session': seances_data.session,
                'dates': seances_data.dates,
                'total_sessions': sum(len(sessions) for sessions in seances_data.date_seances.values())
            }
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def import_seances_from_file_content(file_content, filename):
    """Import seances data from CSV or XLSX file content, replacing all existing data"""
    global seances_data
    
    try:
        # Create temporary file from content
        os.makedirs('data', exist_ok=True)
        
        # Determine file type from filename
        file_ext = os.path.splitext(filename)[1].lower()
        temp_file_path = f'data/temp_import{file_ext}'
        
        # Write content to temporary file
        if file_ext == '.xlsx':
            # For XLSX files, content should be binary
            import base64
            if isinstance(file_content, str):
                # If it's base64 encoded, decode it
                try:
                    file_content = base64.b64decode(file_content)
                except:
                    pass
            with open(temp_file_path, 'wb') as f:
                f.write(file_content)
        else:
            # For CSV files, content is text
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                f.write(file_content)
        
        # Validate file format by reading first few lines/rows
        if file_ext == '.csv':
            with open(temp_file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader)
                
                # Check if it has the required columns
                required_columns = ['dateExam', 'h_debut', 'h_fin', 'session', 'type ex', 'semestre', 'enseignant', 'cod_salle']
                if not all(col in header for col in required_columns):
                    return {
                        'success': False, 
                        'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
                    }
        elif file_ext == '.xlsx':
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(temp_file_path)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1] if cell.value is not None]
                workbook.close()
                
                # Clean headers - remove extra whitespace and convert to string
                header = [str(h).strip() for h in header if h is not None]
                
                # Check if it has the required columns
                required_columns = ['dateExam', 'h_debut', 'h_fin', 'session', 'type ex', 'semestre', 'enseignant', 'cod_salle']
                missing_columns = [col for col in required_columns if col not in header]
                
                if missing_columns:
                    return {
                        'success': False, 
                        'error': f'Invalid XLSX format. Missing columns: {", ".join(missing_columns)}. Found columns: {", ".join(header)}'
                    }
            except ImportError:
                return {
                    'success': False, 
                    'error': 'openpyxl library is required for Excel files. Please install it.'
                }
            except Exception as e:
                return {
                    'success': False, 
                    'error': f'Error reading XLSX file: {str(e)}'
                }
        else:
            return {
                'success': False, 
                'error': f'Unsupported file format: {file_ext}. Only .csv and .xlsx are supported.'
            }
        
        # Load seances from temporary file, replacing all existing data
        seances_data = Seances.from_csv(temp_file_path)
        
        # Clean up temporary file
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        # Save the imported state
        save_current_state()
        
        return {
            'success': True,
            'message': 'Seances data imported successfully',
            'seances_info': {
                'semester': seances_data.semester,
                'exam_type': seances_data.exam_type,
                'session': seances_data.session,
                'dates': seances_data.dates,
                'total_sessions': sum(len(sessions) for sessions in seances_data.date_seances.values())
            }
        }
    except Exception as e:
        # Clean up temporary file in case of error
        temp_file_path = f'data/temp_import{file_ext if "file_ext" in locals() else ".csv"}'
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def import_seances_from_csv_content(csv_content):
    """Import seances data from CSV content, replacing all existing data"""
    global seances_data
    
    try:
        # Create temporary CSV file from content
        os.makedirs('data', exist_ok=True)
        temp_csv_path = 'data/temp_import.csv'
        
        # Write content to temporary file
        with open(temp_csv_path, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Validate CSV format by reading first few lines
        with open(temp_csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            
            # Check if it has the required columns
            required_columns = ['dateExam', 'h_debut', 'h_fin', 'session', 'type ex', 'semestre', 'enseignant', 'cod_salle']
            if not all(col in header for col in required_columns):
                return {
                    'success': False, 
                    'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
                }
        
        # Load seances from temporary CSV file, replacing all existing data
        seances_data = Seances.from_csv(temp_csv_path)
        
        # Clean up temporary file
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        # Save the imported state
        save_current_state()
        
        return {
            'success': True,
            'message': 'Seances data imported successfully',
            'seances_info': {
                'semester': seances_data.semester,
                'exam_type': seances_data.exam_type,
                'session': seances_data.session,
                'dates': seances_data.dates,
                'total_sessions': sum(len(sessions) for sessions in seances_data.date_seances.values())
            }
        }
    except Exception as e:
        # Clean up temporary file in case of error
        temp_csv_path = 'data/temp_import.csv'
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        return {
            'success': False,
            'error': str(e)
        }

def detect_teacher_conflicts():
    """Detect conflicts between teacher assignments and their unavailability preferences"""
    global enseignants_data, seances_data
    
    if not enseignants_data or not seances_data:
        return {}
    
    conflicts = {}
    
    # Get day-seance to teacher mapping
    day_seance_teachers = seances_data.get_day_seance_teachers_mapping()
    
    # For each teacher with souhaits, check if they're assigned to sessions they're unavailable for
    for enseignant in enseignants_data.enseignants_list:
        if not enseignant.souhaits or not enseignant.participe_surveillance:
            continue
            
        teacher_conflicts = []
        
        # Check each day-seance combination
        for (day_index, seance_index), teacher_codes in day_seance_teachers.items():
            # Check if this teacher is assigned to this session
            if enseignant.code and enseignant.code in teacher_codes:
                # Check if teacher is unavailable for this day-session
                # day_index is 0-based, but souhaits use 1-based day numbers
                day_number = day_index + 1
                seance_number = seance_index + 1
                
                if not enseignant.is_available(day_number, seance_number):
                    # Get the actual date for this day_index
                    date = seances_data.dates[day_index] if day_index < len(seances_data.dates) else f"Day {day_number}"
                    teacher_conflicts.append({
                        'day_index': day_index,
                        'seance_index': seance_index,
                        'day_number': day_number,
                        'seance_number': seance_number,
                        'date': date,
                        'seance_name': f"S{seance_number}"
                    })
        
        if teacher_conflicts:
            conflicts[enseignant.email] = teacher_conflicts
    
    return conflicts

@eel.expose
def get_enseignants_list():
    """Get list of all teachers"""
    global enseignants_data
    
    if not enseignants_data:
        return []
    
    # Detect conflicts for all teachers
    conflicts = detect_teacher_conflicts()
        
    return [
        {
            'code': ens.code,
            'nom': ens.nom,
            'prenom': ens.prenom,
            'email': ens.email,
            'grade': ens.grade,
            'participe_surveillance': ens.participe_surveillance,
            'has_souhaits': ens.souhaits is not None,
            'has_conflicts': ens.email in conflicts,
            'conflicts': conflicts.get(ens.email, [])
        }
        for ens in enseignants_data.enseignants_list
    ]

@eel.expose
def get_conflict_summary():
    """Get summary of all conflicts between teacher assignments and preferences"""
    global enseignants_data
    
    if not enseignants_data:
        return {'has_conflicts': False, 'conflicts': [], 'total_conflicts': 0}
    
    conflicts = detect_teacher_conflicts()
    
    if not conflicts:
        return {'has_conflicts': False, 'conflicts': [], 'total_conflicts': 0}
    
    # Format conflicts for display
    conflict_list = []
    total_conflict_count = 0
    
    for email, teacher_conflicts in conflicts.items():
        enseignant = enseignants_data.get_enseignant_by_email(email)
        if enseignant:
            conflict_list.append({
                'teacher_name': f"{enseignant.prenom} {enseignant.nom}",
                'email': email,
                'conflicts': teacher_conflicts
            })
            total_conflict_count += len(teacher_conflicts)
    
    return {
        'has_conflicts': True,
        'conflicts': conflict_list,
        'total_conflicts': total_conflict_count
    }

@eel.expose
def get_teacher_assignments(email):
    """Get all assignments for a specific teacher"""
    global enseignants_data, seances_data, assignements_data
    
    if not enseignants_data or not seances_data:
        return {'success': False, 'assignments': []}
    
    # Find the teacher
    enseignant = enseignants_data.get_enseignant_by_email(email)
    if not enseignant or not enseignant.code:
        return {'success': False, 'assignments': []}
    
    assignments = []
    
    # Check if we have assignments data, prefer that over seances data
    if assignements_data and hasattr(assignements_data, 'assignments'):
        # Use assignment system data
        for (day, seance), teacher_ids in assignements_data.assignments.items():
            if enseignant.code in teacher_ids:
                day_index = day - 1
                seance_index = seance - 1
                date = seances_data.dates[day_index] if day_index < len(seances_data.dates) else f"Day {day}"
                
                # Get seance details
                seance_obj = None
                if (date in seances_data.date_seances and 
                    seance_index < len(seances_data.date_seances[date])):
                    seance_obj = seances_data.date_seances[date][seance_index]
                
                h_debut = seance_obj.h_debut if seance_obj else "N/A"
                h_fin = seance_obj.h_fin if seance_obj else "N/A"
                salles = list(seance_obj.salles) if seance_obj else []
                
                assignments.append({
                    'day_index': day_index,
                    'seance_index': seance_index,
                    'day_number': day,
                    'seance_number': seance,
                    'date': date,
                    'seance_name': f"S{seance}",
                    'h_debut': h_debut,
                    'h_fin': h_fin,
                    'rooms': salles
                })
    else:
        # Fallback to seances data
        day_seance_teachers = seances_data.get_day_seance_teachers_mapping()
        
        # Find all assignments for this teacher
        for (day_index, seance_index), teacher_codes in day_seance_teachers.items():
            if enseignant.code in teacher_codes:
                day_number = day_index + 1
                seance_number = seance_index + 1
                date = seances_data.dates[day_index] if day_index < len(seances_data.dates) else f"Day {day_number}"
                
                # Get seance details
                seance_obj = None
                if (date in seances_data.date_seances and 
                    seance_index < len(seances_data.date_seances[date])):
                    seance_obj = seances_data.date_seances[date][seance_index]
                
                h_debut = seance_obj.h_debut if seance_obj else "N/A"
                h_fin = seance_obj.h_fin if seance_obj else "N/A"
                salles = list(seance_obj.salles) if seance_obj else []
                
                assignments.append({
                    'day_index': day_index,
                    'seance_index': seance_index,
                    'day_number': day_number,
                    'seance_number': seance_number,
                    'date': date,
                    'seance_name': f"S{seance_number}",
                    'h_debut': h_debut,
                    'h_fin': h_fin,
                    'rooms': salles
                })
    
    return {
        'success': True,
        'assignments': assignments
    }

@eel.expose
def get_teacher_responsibilities(email):
    """Get all seances where a teacher is responsible (not just assigned for surveillance)"""
    global enseignants_data, seances_data
    
    if not enseignants_data or not seances_data:
        return {'success': False, 'responsibilities': []}
    
    # Find the teacher
    enseignant = enseignants_data.get_enseignant_by_email(email)
    if not enseignant or not enseignant.code:
        return {'success': False, 'responsibilities': []}
    
    responsibilities = []
    
    # Check each date and seance to see if teacher is responsible
    for day_index, date in enumerate(seances_data.dates):
        if date in seances_data.date_seances:
            for seance_index, seance in enumerate(seances_data.date_seances[date]):
                # Check if teacher's code is in the seance's enseignants (responsible teachers)
                if enseignant.code in seance.responsables:
                    day_number = day_index + 1
                    seance_number = seance_index + 1
                    
                    responsibilities.append({
                        'day_index': day_index,
                        'seance_index': seance_index,
                        'day_number': day_number,
                        'seance_number': seance_number,
                        'date': date,
                        'seance_name': f"S{seance_number}",
                        'h_debut': seance.h_debut,
                        'h_fin': seance.h_fin,
                        'rooms': list(seance.salles)
                    })
    
    return {
        'success': True,
        'responsibilities': responsibilities
    }

@eel.expose
def get_all_teachers_with_assignments():
    """Get all teachers with their assignment counts and details"""
    global enseignants_data, seances_data, assignements_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'teachers': []}
        
        teachers_with_assignments = []
        
        for teacher in enseignants_data.enseignants_list:
            # Get assignments for this teacher
            assignment_result = get_teacher_assignments(teacher.email)
            assignments = assignment_result.get('assignments', []) if assignment_result.get('success') else []
            
            # Calculate statistics
            total_assignments = len(assignments)
            unique_dates = len(set(assignment.get('date') for assignment in assignments))
            
            # Get conflicts if available
            conflicts = teacher.conflicts if hasattr(teacher, 'conflicts') and teacher.conflicts else []
            
            teachers_with_assignments.append({
                'email': teacher.email,
                'nom': teacher.nom,
                'prenom': teacher.prenom,
                'grade': teacher.grade,
                'code': teacher.code,
                'participe_surveillance': teacher.participe_surveillance,
                'has_souhaits': teacher.souhaits is not None,
                'assignment_count': total_assignments,
                'unique_dates': unique_dates,
                'conflicts_count': len(conflicts),
                'assignments': assignments
            })
        
        return {
            'success': True,
            'teachers': teachers_with_assignments
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def toggle_teacher_assignment_for_seance(teacher_email, day, seance, assign=True, force_unavailable=False):
    """Toggle a teacher's assignment to a specific seance"""
    global enseignants_data, assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        # Find the teacher
        teacher = enseignants_data.get_enseignant_by_email(teacher_email)
        if not teacher or not teacher.code:
            return {
                'success': False,
                'error': 'Teacher not found or has no code'
            }
        
        teacher_id = teacher.code
        
        if assign:
            # Use the new assignment method
            result = assignements_data.assign_teacher_to_seance(day, seance, teacher_id, force_unavailable)
            if result['success']:
                save_current_state()
                return {
                    'success': True,
                    'message': f'Teacher {result["teacher_name"]} assigned to Day {day}, S{seance}',
                    'is_conflict': result.get('is_conflict', False)
                }
            else:
                return result
        else:
            # Remove teacher from seance
            success = assignements_data.remove_teacher_from_seance(day, seance, teacher_id)
            if success:
                save_current_state()
                return {
                    'success': True,
                    'message': f'Teacher {teacher.prenom} {teacher.nom} removed from Day {day}, S{seance}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Teacher is not assigned to this seance'
                }
    
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_all_assignment_conflicts():
    """Get all assignment conflicts (teachers assigned to unavailable seances)"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        conflicts = assignements_data.get_all_conflicts()
        return {
            'success': True,
            'conflicts': conflicts
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def check_assignment_conflict(day, seance, teacher_id):
    """Check if a specific assignment is a conflict"""
    global assignements_data
    
    try:
        if not assignements_data:
            return False
        
        return assignements_data.is_assignment_conflict(day, seance, teacher_id)
    
    except Exception as e:
        return False

@eel.expose
def add_enseignant(nom, prenom, email, grade, code=None, participe_surveillance=False):
    """Add a new teacher"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            enseignants_data = Enseignants()
        
        # Validate email uniqueness
        existing = enseignants_data.get_enseignant_by_email(email)
        if existing:
            return {'success': False, 'error': f'Email {email} already exists'}
        
        # Convert code to int if provided and valid
        parsed_code = None
        if code and str(code).strip():
            try:
                parsed_code = int(code)
                # Check for code uniqueness
                existing_code = enseignants_data.get_enseignant_by_code(parsed_code)
                if existing_code:
                    return {'success': False, 'error': f'Code {parsed_code} already exists'}
            except ValueError:
                return {'success': False, 'error': 'Invalid code format'}
        
        from enseignants import Enseignant
        new_enseignant = Enseignant(
            nom=nom.strip(),
            prenom=prenom.strip(),
            email=email.strip(),
            grade=grade.strip(),
            code=parsed_code,
            participe_surveillance=bool(participe_surveillance)
        )
        
        enseignants_data.add_enseignant(new_enseignant)
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': f'Teacher {prenom} {nom} added successfully'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def update_enseignant(original_email, nom, prenom, email, grade, code=None, participe_surveillance=False):
    """Update an existing teacher"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available'}
        
        # Find the teacher by original email
        enseignant = enseignants_data.get_enseignant_by_email(original_email)
        if not enseignant:
            return {'success': False, 'error': 'Teacher not found'}
        
        # If email is changing, check for uniqueness
        if email != original_email:
            existing = enseignants_data.get_enseignant_by_email(email)
            if existing:
                return {'success': False, 'error': f'Email {email} already exists'}
        
        # Convert and validate code
        parsed_code = None
        if code and str(code).strip():
            try:
                parsed_code = int(code)
                # Check for code uniqueness (excluding current teacher)
                existing_code = enseignants_data.get_enseignant_by_code(parsed_code)
                if existing_code and existing_code.email != original_email:
                    return {'success': False, 'error': f'Code {parsed_code} already exists'}
            except ValueError:
                return {'success': False, 'error': 'Invalid code format'}
        
        # Update the teacher
        enseignant.nom = nom.strip()
        enseignant.prenom = prenom.strip()
        enseignant.email = email.strip()
        enseignant.grade = grade.strip()
        enseignant.code = parsed_code
        enseignant.participe_surveillance = bool(participe_surveillance)
        
        # Update unique grades
        enseignants_data.unique_grades.add(grade.strip())
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': f'Teacher {prenom} {nom} updated successfully'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def delete_enseignant(email):
    """Delete a teacher by email"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available'}
        
        # Find the teacher
        enseignant = enseignants_data.get_enseignant_by_email(email)
        if not enseignant:
            return {'success': False, 'error': 'Teacher not found'}
        
        # Remove from list
        enseignants_data.enseignants_list.remove(enseignant)
        
        # Update unique grades
        enseignants_data.unique_grades = set(ens.grade for ens in enseignants_data.enseignants_list)
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': f'Teacher {enseignant.prenom} {enseignant.nom} deleted successfully'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def import_enseignants_from_csv():
    """Import teachers from CSV file, replacing existing data"""
    global enseignants_data
    
    try:
        csv_path = 'data/enseignants.csv'
        if not os.path.exists(csv_path):
            return {'success': False, 'error': f'CSV file not found: {csv_path}'}
        
        enseignants_data = Enseignants.from_csv(csv_path)
        save_current_state()
        
        return {
            'success': True,
            'message': 'Teachers imported successfully',
            'count': len(enseignants_data.enseignants_list)
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def import_enseignants_from_file_content(file_content, filename):
    """Import teachers from CSV or XLSX file content"""
    global enseignants_data
    
    try:
        # Create temporary file from content
        os.makedirs('data', exist_ok=True)
        
        # Determine file type from filename
        file_ext = os.path.splitext(filename)[1].lower()
        temp_file_path = f'data/temp_enseignants{file_ext}'
        
        # Write content to temporary file
        if file_ext == '.xlsx':
            # For XLSX files, content should be binary
            import base64
            if isinstance(file_content, str):
                # If it's base64 encoded, decode it
                try:
                    file_content = base64.b64decode(file_content)
                except:
                    pass
            with open(temp_file_path, 'wb') as f:
                f.write(file_content)
        else:
            # For CSV files, content is text
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                f.write(file_content)
        
        # Validate file format by reading first few lines/rows
        if file_ext == '.csv':
            from io import StringIO
            csv_reader = csv.reader(StringIO(file_content))
            header = next(csv_reader)
            
            required_columns = ['nom_ens', 'prenom_ens', 'email_ens', 'grade_code_ens', 'code_smartex_ens', 'participe_surveillance']
            if not all(col in header for col in required_columns):
                return {
                    'success': False, 
                    'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
                }
        elif file_ext == '.xlsx':
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(temp_file_path)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1] if cell.value is not None]
                workbook.close()
                
                # Clean headers - remove extra whitespace and convert to string
                header = [str(h).strip() for h in header if h is not None]
                
                required_columns = ['nom_ens', 'prenom_ens', 'email_ens', 'grade_code_ens', 'code_smartex_ens', 'participe_surveillance']
                missing_columns = [col for col in required_columns if col not in header]
                
                if missing_columns:
                    return {
                        'success': False, 
                        'error': f'Invalid XLSX format. Missing columns: {", ".join(missing_columns)}. Found columns: {", ".join(header)}'
                    }
            except ImportError:
                return {
                    'success': False, 
                    'error': 'openpyxl library is required for Excel files. Please install it.'
                }
            except Exception as e:
                return {
                    'success': False, 
                    'error': f'Error reading XLSX file: {str(e)}'
                }
        else:
            return {
                'success': False, 
                'error': f'Unsupported file format: {file_ext}. Only .csv and .xlsx are supported.'
            }
        
        # Load teachers from temporary file
        enseignants_data = Enseignants.from_csv(temp_file_path)
        
        # Clean up temporary file
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': 'Teachers imported successfully',
            'count': len(enseignants_data.enseignants_list)
        }
    except Exception as e:
        # Clean up temporary file in case of error
        temp_file_path = f'data/temp_enseignants{file_ext if "file_ext" in locals() else ".csv"}'
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        return {'success': False, 'error': str(e)}

@eel.expose
def import_enseignants_from_csv_content(csv_content):
    """Import teachers from CSV content"""
    global enseignants_data
    
    try:
        from io import StringIO
        
        # Validate CSV format
        csv_reader = csv.reader(StringIO(csv_content))
        header = next(csv_reader)
        
        required_columns = ['nom_ens', 'prenom_ens', 'email_ens', 'grade_code_ens', 'code_smartex_ens', 'participe_surveillance']
        if not all(col in header for col in required_columns):
            return {
                'success': False, 
                'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
            }
        
        # Create temporary file
        os.makedirs('data', exist_ok=True)
        temp_csv_path = 'data/temp_enseignants.csv'
        
        with open(temp_csv_path, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Load teachers from temporary file
        enseignants_data = Enseignants.from_csv(temp_csv_path)
        
        # Clean up temporary file
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': 'Teachers imported successfully',
            'count': len(enseignants_data.enseignants_list)
        }
    except Exception as e:
        # Clean up temporary file in case of error
        temp_csv_path = 'data/temp_enseignants.csv'
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        return {'success': False, 'error': str(e)}

@eel.expose
def clear_all_seances():
    """Clear all seances and souhaits data"""
    global seances_data, enseignants_data
    
    try:
        # Clear seances data
        seances_data = Seances()
        
        # Clear souhaits from all teachers
        if enseignants_data:
            enseignants_data.clear_all_souhaits()
        
        # Save current state
        save_current_state()
        
        return {
            'success': True,
            'message': 'Toutes les séances et souhaits ont été supprimés'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def clear_all_enseignants():
    """Clear all teachers data"""
    global enseignants_data
    
    try:
        enseignants_data = Enseignants()
        
        # Save to pickle
        save_current_state()
        
        return {
            'success': True,
            'message': 'Tous les enseignants ont été supprimés'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def export_enseignants_csv():
    """Export current enseignants data as CSV content"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data to export'}
        
        from io import StringIO
        
        output = StringIO()
        fieldnames = ['nom_ens', 'prenom_ens', 'email_ens', 'grade_code_ens', 'code_smartex_ens', 'participe_surveillance']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        for ens in enseignants_data.enseignants_list:
            writer.writerow({
                'nom_ens': ens.nom,
                'prenom_ens': ens.prenom,
                'email_ens': ens.email,
                'grade_code_ens': ens.grade,
                'code_smartex_ens': ens.code if ens.code else '',
                'participe_surveillance': 'TRUE' if ens.participe_surveillance else 'FALSE'
            })
        
        csv_content = output.getvalue()
        output.close()
        
        return {
            'success': True,
            'csv_content': csv_content,
            'filename': f'enseignants_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def import_souhaits_from_file_content(file_content, filename):
    """Import teacher preferences from CSV or XLSX file content"""
    global enseignants_data, seances_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available. Please import teachers first.'}
        
        if not seances_data:
            return {'success': False, 'error': 'No seances data available. Please import seances first.'}
        
        # Create temporary file from content
        os.makedirs('data', exist_ok=True)
        
        # Determine file type from filename
        file_ext = os.path.splitext(filename)[1].lower()
        temp_file_path = f'data/temp_souhaits{file_ext}'
        
        # Write content to temporary file
        if file_ext == '.xlsx':
            # For XLSX files, content should be binary
            import base64
            if isinstance(file_content, str):
                # If it's base64 encoded, decode it
                try:
                    file_content = base64.b64decode(file_content)
                except:
                    pass
            with open(temp_file_path, 'wb') as f:
                f.write(file_content)
        else:
            # For CSV files, content is text
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                f.write(file_content)
        
        # Validate file format by reading first few lines/rows
        if file_ext == '.csv':
            from io import StringIO
            csv_reader = csv.reader(StringIO(file_content))
            header = next(csv_reader)
            
            required_columns = ['Enseignant', 'Semestre', 'Session', 'Jour', 'Séances']
            if not all(col in header for col in required_columns):
                return {
                    'success': False, 
                    'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
                }
        elif file_ext == '.xlsx':
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(temp_file_path)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1] if cell.value is not None]
                workbook.close()
                
                # Clean headers - remove extra whitespace and convert to string
                header = [str(h).strip() for h in header if h is not None]
                
                required_columns = ['Enseignant', 'Semestre', 'Session', 'Jour', 'Séances']
                missing_columns = [col for col in required_columns if col not in header]
                
                if missing_columns:
                    return {
                        'success': False, 
                        'error': f'Invalid XLSX format. Missing columns: {", ".join(missing_columns)}. Found columns: {", ".join(header)}'
                    }
            except ImportError:
                return {
                    'success': False, 
                    'error': 'openpyxl library is required for Excel files. Please install it.'
                }
            except Exception as e:
                return {
                    'success': False, 
                    'error': f'Error reading XLSX file: {str(e)}'
                }
        else:
            return {
                'success': False, 
                'error': f'Unsupported file format: {file_ext}. Only .csv and .xlsx are supported.'
            }
        
        # Load souhaits from temporary file (clear existing souhaits first)
        errors = enseignants_data.load_souhaits_from_csv(temp_file_path, seances_data, clear_existing=True)
        
        # Clean up temporary file
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        # Save updated enseignants data
        save_current_state()
        
        result = {'success': True, 'message': 'Souhaits imported successfully'}
        if errors:
            result['warnings'] = errors
            
        return result
        
    except Exception as e:
        # Clean up temporary file in case of error
        temp_file_path = f'data/temp_souhaits{file_ext if "file_ext" in locals() else ".csv"}'
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        return {'success': False, 'error': str(e)}

@eel.expose
def import_souhaits_from_csv_content(csv_content):
    """Import teacher preferences from CSV content"""
    global enseignants_data, seances_data
    
    try:
        import csv
        from io import StringIO
        
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available. Please import teachers first.'}
        
        if not seances_data:
            return {'success': False, 'error': 'No seances data available. Please import seances first.'}
        
        # Validate CSV format
        csv_reader = csv.reader(StringIO(csv_content))
        header = next(csv_reader)
        
        required_columns = ['Enseignant', 'Semestre', 'Session', 'Jour', 'Séances']
        if not all(col in header for col in required_columns):
            return {
                'success': False, 
                'error': f'Invalid CSV format. Required columns: {", ".join(required_columns)}'
            }
        
        # Create temporary file
        os.makedirs('data', exist_ok=True)
        temp_csv_path = 'data/temp_souhaits.csv'
        
        with open(temp_csv_path, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Load souhaits from temporary file (clear existing souhaits first)
        errors = enseignants_data.load_souhaits_from_csv(temp_csv_path, seances_data, clear_existing=True)
        
        # Clean up temporary file
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        # Save updated enseignants data
        save_current_state()
        
        result = {'success': True, 'message': 'Souhaits imported successfully'}
        if errors:
            result['warnings'] = errors
            
        return result
        
    except Exception as e:
        # Clean up temporary file in case of error
        temp_csv_path = 'data/temp_souhaits.csv'
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        return {'success': False, 'error': str(e)}

@eel.expose
def get_enseignant_souhaits(email):
    """Get souhaits for a specific teacher"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available'}
        
        enseignant = enseignants_data.get_enseignant_by_email(email)
        if not enseignant:
            return {'success': False, 'error': 'Teacher not found'}
        
        if not enseignant.souhaits:
            return {
                'success': True,
                'souhaits': {
                    'semestre': None,
                    'session': None,
                    'unavailable_slots': []
                }
            }
        
        return {
            'success': True,
            'souhaits': {
                'semestre': enseignant.souhaits.semestre,
                'session': enseignant.souhaits.session,
                'unavailable_slots': list(enseignant.souhaits.unavailable_slots)
            }
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def update_enseignant_souhaits(email, semestre, session, unavailable_slots):
    """Update souhaits for a specific teacher"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available'}
        
        enseignant = enseignants_data.get_enseignant_by_email(email)
        if not enseignant:
            return {'success': False, 'error': 'Teacher not found'}
        
        from enseignants import Souhaits
        
        # Create or update souhaits
        if not enseignant.souhaits:
            enseignant.souhaits = Souhaits(semestre=semestre, session=session)
        else:
            enseignant.souhaits.semestre = semestre
            enseignant.souhaits.session = session
            enseignant.souhaits.unavailable_slots.clear()
        
        # Add unavailable slots
        for slot in unavailable_slots:
            if isinstance(slot, list) and len(slot) == 2:
                enseignant.souhaits.add_unavailable_slot(int(slot[0]), int(slot[1]))
        
        # Save updated data
        save_current_state()
        
        return {
            'success': True,
            'message': f'Souhaits updated for {enseignant.prenom} {enseignant.nom}'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def get_exam_schedule_for_souhaits():
    """Get exam schedule information for souhaits calendar"""
    global seances_data
    
    try:
        if not seances_data or not seances_data.dates:
            return {
                'success': True,
                'dates': [],
                'sessions_by_date': {},
                'warning': 'No exam sessions available. Please import seances data first.'
            }
        
        # Create a mapping of dates to their sessions
        sessions_by_date = {}
        for date in seances_data.dates:
            if date in seances_data.date_seances:
                sessions = []
                for i, seance in enumerate(seances_data.date_seances[date]):
                    sessions.append({
                        'index': i + 1,  # Session number (1-based)
                        'h_debut': seance.h_debut,
                        'h_fin': seance.h_fin,
                        'name': f'S{i + 1}'
                    })
                sessions_by_date[date] = sessions
            else:
                sessions_by_date[date] = []
        
        return {
            'success': True,
            'dates': seances_data.dates,
            'sessions_by_date': sessions_by_date,
            'semester': seances_data.semester,
            'session': seances_data.session,
            'exam_type': seances_data.exam_type
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def clear_all_souhaits():
    """Clear all souhaits for all teachers"""
    global enseignants_data
    
    try:
        if not enseignants_data:
            return {'success': False, 'error': 'No teachers data available'}
        
        enseignants_data.clear_all_souhaits()
        
        # Save updated data
        save_current_state()
        
        return {
            'success': True,
            'message': 'Tous les souhaits ont été supprimés'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@eel.expose
def get_unique_grades():
    """Get list of unique grades"""
    global enseignants_data
    
    if not enseignants_data:
        return []
    
    return sorted(list(enseignants_data.unique_grades))

# Configuration endpoints
@eel.expose
def get_configuration():
    """Get current configuration"""
    global configuration_data, enseignants_data
    
    try:
        # Initialize configuration if not available
        if not configuration_data:
            configuration_data = load_configuration_state()
        
        # Initialize enseignants if not available
        if not enseignants_data:
            enseignants_original_path = 'data/enseignants.csv'
            if os.path.exists(enseignants_original_path):
                enseignants_data = Enseignants.from_csv(enseignants_original_path)
            else:
                enseignants_data = Enseignants()
        
        # Get configuration summary with validation
        if enseignants_data:
            summary = configuration_data.get_configuration_summary(enseignants_data)
        else:
            summary = {
                'total_grades_in_system': 0,
                'total_configured_grades': len(configuration_data.grade_hours),
                'is_fully_configured': True,
                'missing_grades': [],
                'extra_grades': list(configuration_data.get_all_configured_grades()),
                'configured_grades': dict(configuration_data.grade_hours),
                'total_hours_configured': sum(configuration_data.grade_hours.values()),
                'teachers_per_room': configuration_data.teachers_per_room
            }
        
        return {
            'success': True,
            'configuration': summary
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def set_grade_hours(grade, hours):
    """Set hours for a specific grade"""
    global configuration_data
    
    try:
        if not configuration_data:
            configuration_data = Configuration()
        
        configuration_data.set_grade_hours(grade, hours)
        save_current_state()
        
        return {
            'success': True,
            'message': f'Hours for grade {grade} set to {hours}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def remove_grade_configuration(grade):
    """Remove configuration for a specific grade"""
    global configuration_data
    
    try:
        if not configuration_data:
            return {
                'success': False,
                'error': 'No configuration data available'
            }
        
        removed = configuration_data.remove_grade(grade)
        if removed:
            save_current_state()
            return {
                'success': True,
                'message': f'Configuration for grade {grade} removed'
            }
        else:
            return {
                'success': False,
                'error': f'Grade {grade} not found in configuration'
            }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def set_teachers_per_room(teachers_per_room):
    """Set number of teachers per room"""
    global configuration_data
    
    try:
        if not configuration_data:
            configuration_data = Configuration()
        
        configuration_data.set_teachers_per_room(teachers_per_room)
        save_current_state()
        
        return {
            'success': True,
            'message': f'Teachers per room set to {teachers_per_room}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def set_surplus_teachers_per_room(surplus_teachers_per_room):
    """Set number of surplus teachers per room"""
    global configuration_data
    
    try:
        if not configuration_data:
            configuration_data = Configuration()
        
        # Convert to float
        surplus_teachers_per_room = float(surplus_teachers_per_room)
        configuration_data.set_surplus_teachers_per_room(surplus_teachers_per_room)
        save_current_state()
        
        return {
            'success': True,
            'message': f'Surplus teachers per room set to {surplus_teachers_per_room}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_teacher_requirements():
    """Calculate teacher requirements based on current seances and configuration"""
    global configuration_data, seances_data
    
    try:
        if not configuration_data:
            return {
                'success': False,
                'error': 'Configuration not initialized'
            }
        
        if not seances_data:
            return {
                'success': False,
                'error': 'No seances data available'
            }
        
        requirements = configuration_data.calculate_teacher_requirements(seances_data)
        
        # Convert tuple keys to strings for JSON serialization
        requirements_serializable = {
            f"{day}-{seance}": count 
            for (day, seance), count in requirements.items()
        }
        
        total_teachers_needed = sum(requirements.values())
        
        return {
            'success': True,
            'requirements': requirements_serializable,
            'total_teachers_needed': total_teachers_needed,
            'teachers_per_room': configuration_data.teachers_per_room
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_total_rooms():
    """Get total number of rooms across all seances"""
    global seances_data
    
    try:
        if not seances_data:
            return {
                'success': False,
                'error': 'No seances data available',
                'total_rooms': 0
            }
        
        total_rooms = 0
        for date in seances_data.dates:
            date_seances = seances_data.get_seances_by_date(date)
            for seance in date_seances:
                total_rooms += len(seance.salles)
        
        return {
            'success': True,
            'total_rooms': total_rooms
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'total_rooms': 0
        }

@eel.expose
def get_surveillance_statistics():
    """Calculate surveillance statistics: available vs needed"""
    global configuration_data, seances_data, enseignants_data
    
    try:
        # Initialize configuration if not available
        if not configuration_data:
            configuration_data = load_configuration_state()
        
        # Initialize enseignants if not available
        if not enseignants_data:
            enseignants_original_path = 'data/enseignants.csv'
            if os.path.exists(enseignants_original_path):
                enseignants_data = Enseignants.from_csv(enseignants_original_path)
            else:
                enseignants_data = Enseignants()
        
        if not enseignants_data:
            return {
                'success': False,
                'error': 'Enseignants data not available'
            }
        
        # Calculate available surveillance hours by grade
        available_surveillance = {}
        total_available = 0
        
        for grade in enseignants_data.unique_grades:
            # Get teachers of this grade who participate in surveillance
            grade_teachers = [
                teacher for teacher in enseignants_data.get_enseignants_by_grade(grade)
                if teacher.participe_surveillance
            ]
            
            # Get configured surveillance hours for this grade
            surveillance_hours = configuration_data.get_grade_hours(grade)
            
            # Calculate total available surveillance for this grade
            grade_surveillance = len(grade_teachers) * surveillance_hours
            available_surveillance[grade] = {
                'teachers_count': len(grade_teachers),
                'surveillance_hours': surveillance_hours,
                'total_surveillance': grade_surveillance
            }
            total_available += grade_surveillance
        
        # Calculate needed surveillance
        total_needed = 0
        total_needed_with_surplus = 0
        if seances_data:
            requirements = configuration_data.calculate_teacher_requirements(seances_data)
            total_needed = sum(requirements.values())
            
            # Calculate requirements with surplus
            requirements_with_surplus = {}
            for day_idx, date in enumerate(seances_data.dates, 1):
                date_seances = seances_data.get_seances_by_date(date)
                for seance_idx, seance in enumerate(date_seances, 1):
                    num_rooms = len(seance.salles)
                    # Calculate surplus teachers: rooms * teachers_per_room * surplus_ratio, then round up
                    basic_teachers = num_rooms * configuration_data.teachers_per_room
                    surplus_teachers = int(num_rooms * configuration_data.teachers_per_room * configuration_data.surplus_teachers_per_room)
                    required_teachers_with_surplus = basic_teachers + surplus_teachers
                    requirements_with_surplus[(day_idx, seance_idx)] = required_teachers_with_surplus
            
            total_needed_with_surplus = sum(requirements_with_surplus.values())
        
        # Calculate surplus surveillance needed
        surplus_needed = total_needed_with_surplus - total_needed
        
        return {
            'success': True,
            'available_surveillance': available_surveillance,
            'total_available': total_available,
            'total_needed': total_needed,
            'total_needed_with_surplus': total_needed_with_surplus,
            'surplus_needed': surplus_needed,
            'balance': total_available - total_needed,
            'balance_with_surplus': total_available - total_needed_with_surplus,
            'is_sufficient': total_available >= total_needed,
            'is_sufficient_with_surplus': total_available >= total_needed_with_surplus
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# Assignments endpoints
@eel.expose
def initialize_assignments():
    """Initialize assignments system"""
    global assignements_data, enseignants_data, seances_data, configuration_data
    
    try:
        if not all([enseignants_data, seances_data, configuration_data]):
            return {
                'success': False,
                'error': 'Required data not loaded. Please load seances, enseignants, and configuration first.'
            }
        
        # Check if we already have saved assignments data that's compatible
        if (assignements_data and 
            hasattr(assignements_data, 'enseignants') and 
            hasattr(assignements_data, 'seances') and 
            hasattr(assignements_data, 'configuration') and
            hasattr(assignements_data, 'assignments')):
            
            # Update the references in case the data objects changed
            assignements_data.enseignants = enseignants_data
            assignements_data.seances = seances_data
            assignements_data.configuration = configuration_data
            
            # Reinitialize requirements in case seances changed, preserving existing assignments
            try:
                assignements_data._initialize_requirements()
                # The modified _initialize_assignments will preserve existing assignments
                assignements_data._initialize_assignments()
            except Exception as e:
                print(f"Warning: Error during reinitialization: {e}")
        else:
            # Create new assignments instance only if we don't have existing data
            assignements_data = Assignements(enseignants_data, seances_data, configuration_data)
        
        return {
            'success': True,
            'message': 'Assignments system initialized successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_assignments_overview():
    """Get overview of all assignments"""
    global assignements_data, seances_data, configuration_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        summary = assignements_data.get_assignment_summary()
        
        # Format seance details for frontend
        seance_details = []
        for detail in summary['seance_details']:
            # Get teacher names and check for conflicts
            teacher_names = []
            has_conflicts = False
            for teacher_id in detail['teachers']:
                teacher = enseignants_data.get_enseignant_by_code(teacher_id)
                if teacher:
                    is_conflict = assignements_data.is_assignment_conflict(detail['day'], detail['seance'], teacher_id)
                    if is_conflict:
                        has_conflicts = True
                    
                    teacher_names.append({
                        'id': teacher_id,
                        'name': f"{teacher.prenom} {teacher.nom}",
                        'grade': teacher.grade,
                        'is_conflict': is_conflict
                    })
            
            # Calculate required teachers based on rooms and configuration
            required_teachers = detail['required']  # Default fallback
            if seances_data and configuration_data:
                # Get the actual seance to count rooms
                day_index = detail['day'] - 1  # Convert to 0-based index
                seance_index = detail['seance'] - 1  # Convert to 0-based index
                
                if (day_index < len(seances_data.dates) and 
                    day_index >= 0):
                    date = seances_data.dates[day_index]
                    if (date in seances_data.date_seances and 
                        seance_index < len(seances_data.date_seances[date]) and 
                        seance_index >= 0):
                        seance = seances_data.date_seances[date][seance_index]
                        num_rooms = len(seance.salles) if seance.salles else 1  # At least 1 room
                        teachers_per_room = configuration_data.teachers_per_room if configuration_data.teachers_per_room else 2  # Default to 2
                        required_teachers = num_rooms * teachers_per_room
            
            seance_details.append({
                'day': detail['day'],
                'seance': detail['seance'],
                'required': required_teachers,
                'assigned': detail['assigned'],
                'teachers': teacher_names,
                'is_complete': detail['assigned'] >= required_teachers,
                'is_over_assigned': detail.get('is_over_assigned', False),
                'has_conflicts': has_conflicts
            })
        
        return {
            'success': True,
            'overview': {
                'total_seances': summary['total_seances'],
                'total_requirements': summary['total_requirements'],
                'total_assignments': summary['total_assignments'],
                'fully_satisfied_seances': summary['fully_satisfied_seances'],
                'partially_satisfied_seances': summary['partially_satisfied_seances'],
                'unsatisfied_seances': summary['unsatisfied_seances']
            },
            'seance_details': seance_details
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_seance_assignments(day, seance):
    """Get assignments for a specific seance"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        seance_key = (day, seance)
        if seance_key not in assignements_data.assignments:
            return {
                'success': False,
                'error': f'Seance Day {day}, S{seance} not found'
            }
        
        teacher_ids = assignements_data.assignments[seance_key]
        required = assignements_data.requirements[seance_key]
        
        # Get teacher details
        assigned_teachers = []
        for teacher_id in teacher_ids:
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            if teacher:
                assigned_teachers.append({
                    'id': teacher_id,
                    'code': teacher.code,
                    'name': f"{teacher.prenom} {teacher.nom}",
                    'grade': teacher.grade,
                    'email': teacher.email
                })
        
        return {
            'success': True,
            'day': day,
            'seance': seance,
            'required': required,
            'assigned': len(teacher_ids),
            'teachers': assigned_teachers,
            'is_complete': len(teacher_ids) >= required
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def assign_teacher_to_seance_manual(day, seance, teacher_id):
    """Manually assign a teacher to a seance"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        success = assignements_data.assign_teacher_to_seance(day, seance, teacher_id)
        
        if success:
            save_current_state()  # Save state after assignment
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            teacher_name = f"{teacher.prenom} {teacher.nom}" if teacher else f"Teacher {teacher_id}"
            
            return {
                'success': True,
                'message': f'{teacher_name} assigned to Day {day}, S{seance} successfully'
            }
        else:
            # Get specific error reason
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            if not teacher:
                error_msg = 'Teacher not found'
            elif not teacher.participe_surveillance:
                error_msg = 'Teacher does not participate in surveillance'
            elif not teacher.is_available(day, seance):
                error_msg = 'Teacher is not available for this time slot'
            elif assignements_data.get_teacher_total_surveillances(teacher_id) >= assignements_data.get_teacher_quota(teacher_id):
                quota = assignements_data.get_teacher_quota(teacher_id)
                error_msg = f'Teacher has reached their quota ({quota} hours)'
            elif teacher_id in assignements_data.assignments.get((day, seance), []):
                error_msg = 'Teacher is already assigned to this seance'
            else:
                seance_key = (day, seance)
                if seance_key in assignements_data.requirements:
                    current = len(assignements_data.assignments[seance_key])
                    required = assignements_data.requirements[seance_key]
                    if current >= required:
                        error_msg = 'Seance is already fully staffed'
                    else:
                        error_msg = 'Assignment failed for unknown reason'
                else:
                    error_msg = 'Invalid seance'
            
            return {
                'success': False,
                'error': error_msg
            }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def remove_teacher_from_seance_manual(day, seance, teacher_id):
    """Remove a teacher from a seance"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        success = assignements_data.remove_teacher_from_seance(day, seance, teacher_id)
        
        if success:
            save_current_state()  # Save state after removal
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            teacher_name = f"{teacher.prenom} {teacher.nom}" if teacher else f"Teacher {teacher_id}"
            
            return {
                'success': True,
                'message': f'{teacher_name} removed from Day {day}, S{seance} successfully'
            }
        else:
            return {
                'success': False,
                'error': 'Teacher was not assigned to this seance'
            }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def auto_assign_teachers():
    """Run the automatic teacher assignment algorithm"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        results = assignements_data.auto_assign_teachers()
        
        if results['status'] in ['complete_success', 'partial_success']:
            save_current_state()  # Save state after auto-assignment
            
            # Convert recommendations to strings for frontend
            recommendations_strings = []
            for rec in results.get('recommendations', []):
                recommendations_strings.append({
                    'type': rec.type.value,
                    'message': rec.to_string(),
                    'day': rec.day,
                    'seance': rec.seance,
                    'grade': rec.grade,
                    'teacher_name': rec.teacher_name,
                    'current_quota': rec.current_quota,
                    'suggested_quota': rec.suggested_quota
                })
            
            # Build response based on the enhanced result structure
            response = {
                'success': True,
                'status': results['status'],
                'assignments_made': results['assignments_made'],
                'recommendations': recommendations_strings,
                'message': results.get('message', f'Auto-assignment completed. {results["assignments_made"]} assignments made.'),
                'solver_status': results.get('solver_status', 'UNKNOWN')
            }
            
            # Add detailed statistics if available
            if 'satisfied_seances' in results:
                response['satisfied_seances'] = results['satisfied_seances']
            if 'unsatisfied_seances' in results:
                response['unsatisfied_seances'] = results['unsatisfied_seances']
            if 'total_seances' in results:
                response['total_seances'] = results['total_seances']
            if 'over_assigned_seances' in results:
                response['over_assigned_seances'] = results['over_assigned_seances']
            
            # Legacy support
            if 'unsatisfied_seances' in results:
                response['unsatisfied_seances'] = results['unsatisfied_seances']
            
            return response
        else:
            # Complete failure case
            return {
                'success': False,
                'error': results.get('error', 'Auto-assignment failed'),
                'recommendations': [rec.to_string() for rec in results.get('recommendations', [])]
            }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def clear_all_assignments():
    """Clear all teacher assignments"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        # Clear all assignments
        for seance_key in assignements_data.assignments:
            assignements_data.assignments[seance_key] = []
        
        save_current_state()  # Save state after clearing
        
        return {
            'success': True,
            'message': 'All assignments cleared successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def assign_substitutes():
    """Run the substitute assignment algorithm to improve coverage beyond minimums"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        results = assignements_data.assign_substitutes()
        
        if results['status'] in ['success', 'complete_success']:
            save_current_state()  # Save state after substitute assignment
            
            # Build response
            response = {
                'success': True,
                'status': results['status'],
                'substitutes_assigned': results['substitutes_assigned'],
                'message': results.get('message', f'Substitute assignment completed. {results["substitutes_assigned"]} substitutes assigned.'),
                'solver_status': results.get('solver_status', 'UNKNOWN')
            }
            
            # Add coverage improvements if available
            if 'coverage_improved' in results:
                response['coverage_improved'] = results['coverage_improved']
            
            # Add quota violations if any
            if 'quota_violations' in results:
                response['quota_violations'] = results['quota_violations']
            
            return response
        else:
            return {
                'success': True,  # Still successful even if no improvement
                'status': results['status'],
                'substitutes_assigned': 0,
                'message': results.get('message', 'No substitute assignments possible'),
                'solver_status': results.get('solver_status', 'UNKNOWN')
            }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_available_teachers_for_seance(day, seance):
    """Get list of teachers available for assignment to a specific seance"""
    global assignements_data, enseignants_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        available_teachers = []
        
        for teacher in enseignants_data.get_enseignants_participating_surveillance():
            if teacher.code is None:
                continue
            
            teacher_id = teacher.code
            
            # Check if teacher is already assigned to this seance
            if teacher_id in assignements_data.assignments.get((day, seance), []):
                continue
            
            # Check if teacher is available
            if not teacher.is_available(day, seance):
                continue
            
            # Check if teacher has quota remaining
            current_surveillances = assignements_data.get_teacher_total_surveillances(teacher_id)
            quota = assignements_data.get_teacher_quota(teacher_id)
            
            if current_surveillances >= quota:
                continue
            
            available_teachers.append({
                'id': teacher_id,
                'code': teacher.code,
                'name': f"{teacher.prenom} {teacher.nom}",
                'grade': teacher.grade,
                'email': teacher.email,
                'current_surveillances': current_surveillances,
                'quota': quota,
                'remaining_surveillances': quota - current_surveillances
            })
        
        return {
            'success': True,
            'available_teachers': available_teachers
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def get_all_teachers_for_seance(day, seance):
    """Get list of all teachers with their assignment status for a specific seance"""
    global assignements_data, enseignants_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        all_teachers = []
        assigned_teacher_ids = assignements_data.assignments.get((day, seance), [])
        
        for teacher in enseignants_data.get_enseignants_participating_surveillance():
            if teacher.code is None:
                continue
            
            teacher_id = teacher.code
            is_assigned = teacher_id in assigned_teacher_ids
            
            # Get teacher info
            current_surveillances = assignements_data.get_teacher_total_surveillances(teacher_id)
            quota = assignements_data.get_teacher_quota(teacher_id)
            is_available = teacher.is_available(day, seance)
            has_quota_remaining = current_surveillances < quota
            
            all_teachers.append({
                'id': teacher_id,
                'code': teacher.code,
                'name': f"{teacher.prenom} {teacher.nom}",
                'grade': teacher.grade,
                'email': teacher.email,
                'current_surveillances': current_surveillances,
                'quota': quota,
                'remaining_surveillances': quota - current_surveillances,
                'is_assigned': is_assigned,
                'is_available': is_available,
                'has_quota_remaining': has_quota_remaining,
                'can_be_assigned': is_available and has_quota_remaining and not is_assigned
            })
        
        return {
            'success': True,
            'teachers': all_teachers,
            'seance_info': {
                'day': day,
                'seance': seance,
                'required': assignements_data.requirements.get((day, seance), 0),
                'currently_assigned': len(assigned_teacher_ids)
            }
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def update_seance_assignments(day, seance, assigned_teacher_ids, force_conflicts=False):
    """Update all assignments for a seance (bulk update)"""
    global assignements_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        seance_key = (day, seance)
        if seance_key not in assignements_data.assignments:
            return {
                'success': False,
                'error': f'Seance Day {day}, S{seance} not found'
            }
        
        # Get current assignments
        current_assignments = assignements_data.assignments[seance_key].copy()
        
        # Validate all teacher IDs and check constraints
        changes_made = 0
        errors = []
        conflicts = []
        
        for teacher_id in assigned_teacher_ids:
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            if not teacher:
                errors.append(f"Teacher with ID {teacher_id} not found")
                continue
                
            if not teacher.participe_surveillance:
                errors.append(f"Teacher {teacher.prenom} {teacher.nom} does not participate in surveillance")
                continue
                
            # Check availability - track conflicts but don't block if force_conflicts=True
            if not teacher.is_available(day, seance):
                conflict_info = f"Teacher {teacher.prenom} {teacher.nom} is not available for this time slot"
                if not force_conflicts:
                    errors.append(conflict_info)
                    continue
                else:
                    conflicts.append({
                        'teacher_id': teacher_id,
                        'teacher_name': f"{teacher.prenom} {teacher.nom}",
                        'message': conflict_info
                    })
                
            current_surveillances = assignements_data.get_teacher_total_surveillances(teacher_id)
            quota = assignements_data.get_teacher_quota(teacher_id)
            
            # For teachers being newly assigned, check quota
            if teacher_id not in current_assignments and current_surveillances >= quota:
                errors.append(f"Teacher {teacher.prenom} {teacher.nom} has reached their quota ({current_surveillances}/{quota})")
                continue
        
        if errors:
            return {
                'success': False,
                'error': 'Assignment validation failed: ' + '; '.join(errors)
            }
        
        # Apply changes: remove all current assignments and add new ones
        assignements_data.assignments[seance_key] = assigned_teacher_ids
        
        save_current_state()  # Save state after assignment changes
        
        result = {
            'success': True,
            'message': f'Seance Day {day}, S{seance} assignments updated successfully. {len(assigned_teacher_ids)} teachers assigned.',
            'assigned_count': len(assigned_teacher_ids)
        }
        
        if conflicts:
            result['conflicts'] = conflicts
            result['message'] += f' Warning: {len(conflicts)} conflict(s) detected.'
        
        return result
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@eel.expose
def generate_surveillance_report_for_seance(day, seance, file_name):
    """Generate surveillance report PDF for a specific seance"""
    global assignements_data, enseignants_data, seances_data
    
    try:
        if not assignements_data:
            return {
                'success': False,
                'error': 'Assignments system not initialized'
            }
        
        if not enseignants_data:
            return {
                'success': False,
                'error': 'Teachers data not loaded'
            }
        
        if not seances_data:
            return {
                'success': False,
                'error': 'Sessions data not loaded'
            }
        
        # Get assigned teachers for this seance
        seance_key = (day, seance)
        
        assigned_teacher_ids = assignements_data.assignments.get(seance_key, [])
        
        if not assigned_teacher_ids:
            return {
                'success': False,
                'error': f'No teachers assigned to Day {day}, Session {seance}. Available assignments: {list(assignements_data.assignments.keys())}'
            }
        
        # Get teacher names
        teacher_names = []
        for teacher_id in assigned_teacher_ids:
            teacher = enseignants_data.get_enseignant_by_code(teacher_id)
            if teacher:
                teacher_names.append(f"{teacher.prenom} {teacher.nom}")
        
        if not teacher_names:
            return {
                'success': False,
                'error': f'Could not find teacher information for assigned teacher IDs: {assigned_teacher_ids}'
            }
        
        # Get seance information
        date_str = seances_data.dates[day - 1] if day <= len(seances_data.dates) else f"Day {day}"
        seance_name = f"S{seance}"
        
        # Create output path in web-accessible directory
        os.makedirs('web/generated_reports', exist_ok=True)
        output_path = f"web/generated_reports/{file_name}"
        
        # Generate the PDF
        result = create_surveillance_report(
            enseignants=teacher_names,
            semester=seances_data.semester or "S1",
            exam_type=seances_data.exam_type or "Examen",
            session=seances_data.session or "principale",
            date=date_str,
            seance_name=seance_name,
            output_path=output_path
        )
        
        
        if result.get('status') == 'success':
            # Return web-accessible path and convert to frontend format
            return {
                'success': True,
                'output_path': f"/generated_reports/{file_name}",
                'message': result.get('message', 'PDF généré avec succès'),
                'enseignants_count': len(teacher_names),
                'generated_at': result.get('generated_at')
            }
        else:
            # Convert error format
            return {
                'success': False,
                'error': result.get('message', 'Erreur lors de la génération du PDF')
            }
        
    except Exception as e:
        print(f"Exception in generate_surveillance_report_for_seance: {str(e)}")
        traceback.print_exc()
        return {
            'success': False,
            'error': f'Error generating surveillance report: {str(e)}'
        }

@eel.expose
def generate_all_surveillance_reports(base_name):
    """Generate surveillance reports for all seances with assignments and create a ZIP file"""
    global assignements_data, enseignants_data, seances_data
    
    try:
        if not assignements_data or not enseignants_data or not seances_data:
            return {
                'success': False,
                'error': 'Missing required data (assignments, teachers, or sessions)'
            }
        
        results = []
        generated_count = 0
        temp_files = []  # Keep track of temporary PDF files
        
        # Get all assignments
        all_assignments = assignements_data.assignments
        
        if not all_assignments:
            return {
                'success': False,
                'error': 'No assignments found'
            }
        
        # Create output directory in web-accessible location
        os.makedirs('web/generated_reports', exist_ok=True)
        
        # Create a temporary directory for PDF files
        temp_dir = tempfile.mkdtemp()
        
        try:
            for (day, seance), teacher_ids in all_assignments.items():
                if not teacher_ids:  # Skip empty assignments
                    continue
                    
                # Get teacher names
                teacher_names = []
                for teacher_id in teacher_ids:
                    teacher = enseignants_data.get_enseignant_by_code(teacher_id)
                    if teacher:
                        teacher_names.append(f"{teacher.prenom} {teacher.nom}")
                
                if not teacher_names:
                    continue
                
                # Create filename for this seance
                date_str = seances_data.dates[day - 1] if day <= len(seances_data.dates) else f"Day{day}"
                safe_date = date_str.replace('/', '-').replace('\\', '-')
                filename = f"surveillance_J{day}_S{seance}_{safe_date}.pdf"
                temp_pdf_path = os.path.join(temp_dir, filename)
                
                # Generate the PDF in temporary directory
                result = create_surveillance_report(
                    enseignants=teacher_names,
                    semester=seances_data.semester or "S1",
                    exam_type=seances_data.exam_type or "Examen",
                    session=seances_data.session or "principale",
                    date=date_str,
                    seance_name=f"S{seance}",
                    output_path=temp_pdf_path
                )
                
                results.append({
                    'day': day,
                    'seance': seance,
                    'date': date_str,
                    'filename': filename,
                    'result': result
                })
                
                if result.get('status') == 'success':
                    generated_count += 1
                    temp_files.append((temp_pdf_path, filename))
            
            if generated_count == 0:
                return {
                    'success': False,
                    'error': 'No PDF reports were successfully generated'
                }
            
            # Create ZIP file containing all PDFs
            zip_filename = f"rapports_surveillance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            zip_path = f"web/generated_reports/{zip_filename}"
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for temp_pdf_path, original_filename in temp_files:
                    if os.path.exists(temp_pdf_path):
                        zipf.write(temp_pdf_path, original_filename)
            
            return {
                'success': True,
                'message': f'Generated {generated_count} surveillance reports in ZIP file',
                'generated_count': generated_count,
                'total_attempted': len(results),
                'results': results,
                'zip_download': {
                    'url': f"/generated_reports/{zip_filename}",
                    'filename': zip_filename
                }
            }
            
        finally:
            # Clean up temporary files
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error generating surveillance reports: {str(e)}'
        }

@eel.expose
def generate_teacher_schedule_report(teacher_email, file_name):
    """Generate individual teacher schedule report"""
    global assignements_data, enseignants_data, seances_data
    
    try:
        print(f"Generating teacher schedule for: {teacher_email}, file: {file_name}")
        
        if not assignements_data or not enseignants_data or not seances_data:
            return {
                'success': False,
                'error': 'Missing required data (assignments, teachers, or sessions)'
            }
        
        # Find the teacher
        teacher = enseignants_data.get_enseignant_by_email(teacher_email)
        if not teacher:
            return {
                'success': False,
                'error': 'Teacher not found'
            }
        
        teacher_name = f"{teacher.prenom} {teacher.nom}"
        
        # Get teacher's assignments
        schedule = []
        all_assignments = assignements_data.assignments
        
        for (day, seance), teacher_ids in all_assignments.items():
            if teacher.code in teacher_ids:
                date_str = seances_data.dates[day - 1] if day <= len(seances_data.dates) else f"Day {day}"
                
                # Get seance timing from seances data
                if day <= len(seances_data.dates):
                    date_key = seances_data.dates[day - 1]
                    if date_key in seances_data.date_seances and seance - 1 < len(seances_data.date_seances[date_key]):
                        seance_obj = seances_data.date_seances[date_key][seance - 1]
                        h_debut = seance_obj.h_debut
                        h_fin = seance_obj.h_fin
                    else:
                        h_debut = "08:00:00"
                        h_fin = "12:00:00"
                else:
                    h_debut = "08:00:00"
                    h_fin = "12:00:00"
                
                schedule.append((date_str, h_debut, h_fin))
        
        if not schedule:
            return {
                'success': False,
                'error': f'No assignments found for teacher {teacher_name}'
            }
        
        # Sort schedule by date
        schedule.sort(key=lambda x: x[0])
        
        # Create output path in web-accessible directory
        os.makedirs('web/generated_reports', exist_ok=True)
        output_path = f"web/generated_reports/{file_name}"
        
        # Generate the PDF
        result = create_enseignant_emploi(
            enseignant_name=teacher_name,
            schedule=schedule,
            output_path=output_path
        )
        
        if result.get('status') == 'success':
            # Return web-accessible path and convert to frontend format
            return {
                'success': True,
                'output_path': f"/generated_reports/{file_name}",
                'message': f'Emploi du temps généré avec succès pour {teacher_name}',
                'teacher_name': teacher_name,
                'schedule_count': len(schedule),
                'generated_at': result.get('generated_at')
            }
        else:
            # Convert error format
            return {
                'success': False,
                'error': result.get('message', 'Erreur lors de la génération de l\'emploi du temps')
            }
        
    except Exception as e:
        print(f"Exception in generate_teacher_schedule_report: {str(e)}")
        traceback.print_exc()
        return {
            'success': False,
            'error': f'Error generating teacher schedule: {str(e)}'
        }

@eel.expose
def generate_all_teacher_schedules():
    """Generate all teacher schedules and package them in a ZIP file"""
    global assignements_data, enseignants_data, seances_data
    
    try:
        if not assignements_data or not enseignants_data or not seances_data:
            return {
                'success': False,
                'error': 'Missing required data (assignments, teachers, or sessions)'
            }
        
        results = []
        generated_count = 0
        temp_files = []  # Keep track of temporary PDF files
        
        # Get all assignments
        all_assignments = assignements_data.assignments
        
        if not all_assignments:
            return {
                'success': False,
                'error': 'No assignments found'
            }
        
        # Create output directory in web-accessible location
        os.makedirs('web/generated_reports', exist_ok=True)
        
        # Create a temporary directory for PDF files
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Get all teachers who have assignments
            teachers_with_assignments = set()
            for teacher_ids in all_assignments.values():
                teachers_with_assignments.update(teacher_ids)
            
            for teacher_code in teachers_with_assignments:
                # Get teacher by code
                teacher = enseignants_data.get_enseignant_by_code(teacher_code)
                if not teacher:
                    continue
                
                teacher_name = f"{teacher.prenom} {teacher.nom}"
                
                # Get teacher's assignments
                schedule = []
                
                for (day, seance), teacher_ids in all_assignments.items():
                    if teacher.code in teacher_ids:
                        date_str = seances_data.dates[day - 1] if day <= len(seances_data.dates) else f"Day {day}"
                        
                        # Get seance timing from seances data
                        if day <= len(seances_data.dates):
                            date_key = seances_data.dates[day - 1]
                            if date_key in seances_data.date_seances and seance - 1 < len(seances_data.date_seances[date_key]):
                                seance_obj = seances_data.date_seances[date_key][seance - 1]
                                h_debut = seance_obj.h_debut
                                h_fin = seance_obj.h_fin
                            else:
                                h_debut = "08:00:00"
                                h_fin = "12:00:00"
                        else:
                            h_debut = "08:00:00"
                            h_fin = "12:00:00"
                        
                        schedule.append((date_str, h_debut, h_fin))
                
                if not schedule:
                    continue  # Skip teachers with no assignments
                
                # Sort schedule by date
                schedule.sort(key=lambda x: x[0])
                
                # Create filename for this teacher
                safe_name = f"{teacher.prenom}_{teacher.nom}".replace(' ', '_').replace('/', '_').replace('\\', '_')
                # Remove special characters
                safe_name = ''.join(c for c in safe_name if c.isalnum() or c in ['_', '-'])
                filename = f"emploi_{safe_name}.pdf"
                temp_pdf_path = os.path.join(temp_dir, filename)
                
                # Generate the PDF in temporary directory
                result = create_enseignant_emploi(
                    enseignant_name=teacher_name,
                    schedule=schedule,
                    output_path=temp_pdf_path
                )
                
                results.append({
                    'teacher_name': teacher_name,
                    'teacher_code': teacher.code,
                    'filename': filename,
                    'schedule_count': len(schedule),
                    'result': result
                })
                
                if result.get('status') == 'success':
                    generated_count += 1
                    temp_files.append((temp_pdf_path, filename))
            
            if generated_count == 0:
                return {
                    'success': False,
                    'error': 'No teacher schedules were successfully generated'
                }
            
            # Create ZIP file containing all PDFs
            zip_filename = f"emplois_temps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            zip_path = f"web/generated_reports/{zip_filename}"
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for temp_pdf_path, original_filename in temp_files:
                    if os.path.exists(temp_pdf_path):
                        zipf.write(temp_pdf_path, original_filename)
            
            return {
                'success': True,
                'message': f'Generated {generated_count} teacher schedules in ZIP file',
                'generated_count': generated_count,
                'total_attempted': len(results),
                'results': results,
                'zip_download': {
                    'url': f"/generated_reports/{zip_filename}",
                    'filename': zip_filename
                }
            }
            
        finally:
            # Clean up temporary files
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        
    except Exception as e:
        print(f"Exception in generate_all_teacher_schedules: {str(e)}")
        traceback.print_exc()
        return {
            'success': False,
            'error': f'Error generating teacher schedules: {str(e)}'
        }

def start_app():
    """Start the Eel application"""
    try:
        print("Starting ISI Exams Management Application...")
        
        # First, try to start the server without opening a browser
        port = 8080
        url = f"http://localhost:{port}/seances.html"
        
        print(f"Starting server on port {port}...")
        
        # Try different approaches
        eel.start('seances.html', size=(1200, 800), port=port, mode='default')
                
    except (SystemExit, KeyboardInterrupt):
        print("\nApplication stopped by user")
    except Exception as e:
        print(f"Error starting application: {e}")
        print(f"Please manually open your browser and go to: http://localhost:8080/seances.html")

if __name__ == '__main__':
    start_app()
